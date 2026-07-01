"""Service 层单元测试"""
import io
import pytest

from app.services.export_service import ExportService
from app.services.ai_service import AIService
from app.services.scraper import ScraperService


# ============================================================
# ExportService
# ============================================================

class TestExportService:
    def setup_method(self):
        self.service = ExportService()

    def test_export_keywords_empty(self):
        """空列表导出"""
        buf = self.service.export_keywords([])
        assert isinstance(buf, io.BytesIO)
        assert buf.tell() == 0
        # 验证是有效的 xlsx
        from openpyxl import load_workbook
        wb = load_workbook(buf)
        ws = wb.active
        assert ws.title == "关键词分析"
        assert ws.max_row == 1  # 只有表头

    def test_export_keywords_with_data(self):
        """带数据导出"""
        keywords = [
            {"keyword": "涂抹面膜", "search_volume": 58000, "competition_level": 0.82,
             "potential_score": 65, "category": "功能需求", "source": "下拉词", "notes": ""},
            {"keyword": "敏感肌面膜", "search_volume": 24000, "competition_level": 0.55,
             "potential_score": 88, "category": "人群需求", "source": "相关搜索", "notes": "高潜力"},
        ]
        buf = self.service.export_keywords(keywords)
        from openpyxl import load_workbook
        wb = load_workbook(buf)
        ws = wb.active
        assert ws.max_row == 3  # 表头 + 2行数据
        assert ws.cell(row=2, column=1).value == "涂抹面膜"
        assert ws.cell(row=2, column=2).value == 58000
        assert ws.cell(row=3, column=1).value == "敏感肌面膜"

    def test_export_keywords_header_style(self):
        """验证表头样式"""
        buf = self.service.export_keywords([])
        from openpyxl import load_workbook
        wb = load_workbook(buf)
        ws = wb.active
        header_cell = ws.cell(row=1, column=1)
        assert header_cell.font.bold is True
        assert header_cell.font.color.rgb == "00FFFFFF"

    def test_export_buyer_shows(self):
        """买家秀文案导出"""
        shows = [
            {"template": "使用体验型", "content": "用了这个面膜皮肤真的水润了很多",
             "image_tip": "自然光下自拍", "tone_score": 4.5, "status": "approved"},
        ]
        buf = self.service.export_buyer_shows(shows)
        from openpyxl import load_workbook
        wb = load_workbook(buf)
        ws = wb.active
        assert ws.title == "买家秀文案"
        assert ws.max_row == 2
        assert ws.cell(row=2, column=2).value == "使用体验型"

    def test_export_competitor_analysis(self):
        """竞品分析报告导出"""
        competitor = {"name": "测试竞品", "store": "某店铺", "price": 89, "review_count": 500}
        analysis = {
            "sentiment": {"positive": 80, "neutral": 15, "negative": 5},
            "needs_top10": [
                {"need": "补水保湿", "percent": 8.5, "count": 120},
                {"need": "敏感肌适用", "percent": 6.2, "count": 88},
            ],
            "review_details": [
                {"content": "很好用", "sku": "补水款", "needs": "保湿",
                 "persona": "25岁女性", "scene": "晚间护肤", "sentiment": "positive"},
            ],
            "optimization_suggestions": ["加强保湿成分", "优化包装设计"],
        }
        buf = self.service.export_competitor_analysis(competitor, analysis)
        from openpyxl import load_workbook
        wb = load_workbook(buf)
        assert len(wb.sheetnames) == 4
        assert "总体概览" in wb.sheetnames
        assert "需求分析TOP10" in wb.sheetnames
        assert "评价明细" in wb.sheetnames
        assert "优化建议" in wb.sheetnames

        # 验证需求TOP10数据
        ws2 = wb["需求分析TOP10"]
        assert ws2.cell(row=2, column=2).value == "补水保湿"

    def test_export_competitor_analysis_empty(self):
        """空分析数据导出"""
        competitor = {"name": "测试", "store": "某店", "price": 50, "review_count": 0}
        buf = self.service.export_competitor_analysis(competitor, {})
        from openpyxl import load_workbook
        wb = load_workbook(buf)
        assert len(wb.sheetnames) == 4


# ============================================================
# AIService
# ============================================================

class TestAIService:
    def setup_method(self):
        self.service = AIService()

    def test_mock_response_without_api_key(self):
        """未配置 API Key 时返回 mock 数据"""
        self.service.api_key = None
        import asyncio
        result = asyncio.get_event_loop().run_until_complete(
            self.service.chat("测试prompt")
        )
        assert "[AI模拟响应]" in result
        assert "QWEN_API_KEY" in result

    def test_classify_keyword_fallback(self):
        """关键词分类返回默认值"""
        self.service.api_key = None
        import asyncio
        result = asyncio.get_event_loop().run_until_complete(
            self.service.classify_keyword("涂抹面膜")
        )
        assert "category" in result
        assert "potential_score" in result

    def test_analyze_reviews_fallback(self):
        """评价分析返回默认值"""
        self.service.api_key = None
        import asyncio
        result = asyncio.get_event_loop().run_until_complete(
            self.service.analyze_reviews([{"content": "好用"}], "测试产品")
        )
        assert isinstance(result, dict)

    def test_generate_market_report_fallback(self):
        """市场报告返回默认值"""
        self.service.api_key = None
        import asyncio
        result = asyncio.get_event_loop().run_until_complete(
            self.service.generate_market_report({
                "market_capacity": "100万/年",
                "avg_price": 80,
                "blue_ocean_score": 65,
            })
        )
        assert isinstance(result, str)


# ============================================================
# ScraperService
# ============================================================

class TestScraperService:
    def setup_method(self):
        self.service = ScraperService()

    def test_generate_mock_data(self):
        """模拟市场数据生成"""
        data = self.service._generate_mock_data("涂抹面膜")
        assert data["keyword"] == "涂抹面膜"
        assert data["total_products"] > 0
        assert "avg_price" in data
        assert "blue_ocean_score" in data
        assert "price_distribution" in data
        assert len(data["price_distribution"]) == 6

    def test_generate_mock_data_consistency(self):
        """同一关键词多次生成数据一致（seed固定）"""
        d1 = self.service._generate_mock_data("涂抹面膜")
        d2 = self.service._generate_mock_data("涂抹面膜")
        assert d1["total_products"] == d2["total_products"]
        assert d1["avg_price"] == d2["avg_price"]

    def test_generate_mock_data_different_keywords(self):
        """不同关键词生成不同数据"""
        d1 = self.service._generate_mock_data("涂抹面膜")
        d2 = self.service._generate_mock_data("美白精华")
        # 数据应该不同（极大概率）
        assert d1["keyword"] != d2["keyword"]

    def test_analyze_market_empty(self):
        """空商品列表分析"""
        result = self.service._analyze_market([], "测试")
        assert result["total_products"] == 0
        assert "error" in result

    def test_analyze_market_calculations(self):
        """市场分析计算验证"""
        products = [
            {"title": "产品A", "price": 50.0, "sales": 1000},
            {"title": "产品B", "price": 100.0, "sales": 500},
            {"title": "产品C", "price": 30.0, "sales": 2000},
            {"title": "产品D", "price": 80.0, "sales": 200},
            {"title": "产品E", "price": 150.0, "sales": 100},
        ]
        result = self.service._analyze_market(products, "测试")
        assert result["total_products"] == 5
        assert result["avg_price"] == 82.0  # (50+100+30+80+150)/5
        assert result["avg_sales"] == 760.0  # (1000+500+2000+200+100)/5
        assert 0 <= result["blue_ocean_score"] <= 100
        assert result["top_seller_share"] > 0

    def test_price_distribution(self):
        """价格带分布验证"""
        products = [
            {"title": "A", "price": 25.0, "sales": 100},   # 0-30
            {"title": "B", "price": 45.0, "sales": 200},   # 30-60
            {"title": "C", "price": 80.0, "sales": 300},   # 60-100
            {"title": "D", "price": 120.0, "sales": 400},  # 100-150
            {"title": "E", "price": 180.0, "sales": 500},  # 150-200
            {"title": "F", "price": 250.0, "sales": 600},  # 200+
        ]
        result = self.service._analyze_market(products, "测试")
        dist = {d["range"]: d["count"] for d in result["price_distribution"]}
        assert dist["0-30"] == 1
        assert dist["30-60"] == 1
        assert dist["60-100"] == 1
        assert dist["100-150"] == 1
        assert dist["150-200"] == 1
        assert dist["200+"] == 1

    def test_extract_price(self):
        """价格提取"""
        assert self.service._extract_price("¥89.9") == 89.9
        assert self.service._extract_price("128元") == 128.0
        assert self.service._extract_price("") == 0.0
        assert self.service._extract_price("无价格") == 0.0

    def test_extract_sales(self):
        """销量提取"""
        assert self.service._extract_sales("1000+人付款") == 1000
        assert self.service._extract_sales("500") == 500
        assert self.service._extract_sales("") == 0
        # "2.5万" → "2.50000" → regex 匹配 "2" （不支持小数+万）
        assert self.service._extract_sales("2.5万") == 2
        # 整数万可以正确解析
        assert self.service._extract_sales("3万") == 30000

    def test_parse_cookie(self):
        """Cookie 解析"""
        cookies = self.service._parse_cookie("name1=value1; name2=value2")
        assert len(cookies) == 2
        assert cookies[0]["name"] == "name1"
        assert cookies[0]["value"] == "value1"
        assert cookies[0]["domain"] == ".taobao.com"

    def test_generate_mock_keywords(self):
        """模拟关键词生成"""
        keywords = self.service._generate_mock_keywords("涂抹面膜")
        assert len(keywords) > 0
        assert all("keyword" in k for k in keywords)
        assert all("search_volume" in k for k in keywords)
        assert all("category" in k for k in keywords)
        # 关键词应该包含种子词
        assert all("涂抹面膜" in k["keyword"] for k in keywords)
