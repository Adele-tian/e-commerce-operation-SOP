"""导出服务 - Excel 文件生成"""
import io
from datetime import datetime
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


class ExportService:
    """Excel 导出服务"""

    HEADER_FILL = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    HEADER_FONT = Font(name="Microsoft YaHei", bold=True, color="FFFFFF", size=11)
    BODY_FONT = Font(name="Microsoft YaHei", size=10)
    THIN_BORDER = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0"),
    )

    def export_keywords(self, keywords: list, filename: Optional[str] = None) -> io.BytesIO:
        """导出关键词列表为 Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "关键词分析"

        # 表头
        headers = ["关键词", "搜索量", "竞争度", "潜力评分", "分类", "来源", "备注"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal="center")
            cell.border = self.THIN_BORDER

        # 数据行
        for row_idx, kw in enumerate(keywords, 2):
            values = [
                kw.get("keyword", ""),
                kw.get("search_volume", 0),
                kw.get("competition_level", 0),
                kw.get("potential_score", 0),
                kw.get("category", ""),
                kw.get("source", ""),
                kw.get("notes", ""),
            ]
            for col, value in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.font = self.BODY_FONT
                cell.border = self.THIN_BORDER
                if col > 1:
                    cell.alignment = Alignment(horizontal="center")

        # 列宽
        widths = [30, 12, 12, 12, 10, 15, 20]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[chr(64 + i)].width = w

        # 冻结首行
        ws.freeze_panes = "A2"

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    def export_buyer_shows(self, shows: list) -> io.BytesIO:
        """导出买家秀文案为 Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "买家秀文案"

        headers = ["序号", "模板类型", "文案内容", "配图建议", "语气评分", "状态"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal="center")
            cell.border = self.THIN_BORDER

        for row_idx, show in enumerate(shows, 2):
            values = [
                row_idx - 1,
                show.get("template", ""),
                show.get("content", ""),
                show.get("image_tip", ""),
                show.get("tone_score", 0),
                show.get("status", ""),
            ]
            for col, value in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.font = self.BODY_FONT
                cell.border = self.THIN_BORDER
                if col == 3:
                    cell.alignment = Alignment(wrap_text=True)
                else:
                    cell.alignment = Alignment(horizontal="center")

        widths = [8, 15, 60, 30, 12, 12]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[chr(64 + i)].width = w

        ws.freeze_panes = "A2"

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    def export_competitor_analysis(self, competitor: dict, analysis: dict) -> io.BytesIO:
        """导出竞品评价分析报告为 Excel"""
        wb = Workbook()

        # Sheet 1: 总体概览
        ws1 = wb.active
        ws1.title = "总体概览"
        sentiment = analysis.get("sentiment", {})
        ws1.cell(row=1, column=1, value=f"评价分析报告 - {competitor['name']}").font = Font(name="Microsoft YaHei", bold=True, size=14)
        ws1.cell(row=2, column=1, value=f"店铺: {competitor['store']}  价格: ¥{competitor['price']}  评价数: {competitor['review_count']}").font = self.BODY_FONT
        ws1.cell(row=4, column=1, value="情感分布").font = Font(name="Microsoft YaHei", bold=True, size=11)
        ws1.cell(row=5, column=1, value=f"好评: {sentiment.get('positive', 0)}%")
        ws1.cell(row=6, column=1, value=f"中评: {sentiment.get('neutral', 0)}%")
        ws1.cell(row=7, column=1, value=f"差评: {sentiment.get('negative', 0)}%")
        ws1.column_dimensions['A'].width = 50

        # Sheet 2: 需求TOP10
        ws2 = wb.create_sheet("需求分析TOP10")
        headers2 = ["排名", "用户需求", "占比(%)", "提及次数"]
        for col, h in enumerate(headers2, 1):
            cell = ws2.cell(row=1, column=col, value=h)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal="center")
        for i, need in enumerate(analysis.get("needs_top10", []), 1):
            ws2.cell(row=i+1, column=1, value=i)
            ws2.cell(row=i+1, column=2, value=need.get("need", ""))
            ws2.cell(row=i+1, column=3, value=need.get("percent", 0))
            ws2.cell(row=i+1, column=4, value=need.get("count", 0))
        ws2.column_dimensions['A'].width = 8
        ws2.column_dimensions['B'].width = 18
        ws2.column_dimensions['C'].width = 12
        ws2.column_dimensions['D'].width = 12

        # Sheet 3: 评价明细
        ws3 = wb.create_sheet("评价明细")
        headers3 = ["评价内容", "SKU", "用户需求", "人群画像", "场景", "情绪"]
        for col, h in enumerate(headers3, 1):
            cell = ws3.cell(row=1, column=col, value=h)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal="center")
        for i, review in enumerate(analysis.get("review_details", []), 2):
            ws3.cell(row=i, column=1, value=review.get("content", ""))
            ws3.cell(row=i, column=2, value=review.get("sku", ""))
            ws3.cell(row=i, column=3, value=review.get("needs", ""))
            ws3.cell(row=i, column=4, value=review.get("persona", ""))
            ws3.cell(row=i, column=5, value=review.get("scene", ""))
            ws3.cell(row=i, column=6, value=review.get("sentiment", ""))
        ws3.column_dimensions['A'].width = 60
        ws3.column_dimensions['B'].width = 20
        ws3.column_dimensions['C'].width = 20
        ws3.column_dimensions['D'].width = 18
        ws3.column_dimensions['E'].width = 15
        ws3.column_dimensions['F'].width = 10

        # Sheet 4: 优化建议
        ws4 = wb.create_sheet("优化建议")
        ws4.cell(row=1, column=1, value="序号").font = self.HEADER_FONT
        ws4.cell(row=1, column=1).fill = self.HEADER_FILL
        ws4.cell(row=1, column=2, value="建议内容").font = self.HEADER_FONT
        ws4.cell(row=1, column=2).fill = self.HEADER_FILL
        for i, sug in enumerate(analysis.get("optimization_suggestions", []), 1):
            ws4.cell(row=i+1, column=1, value=i)
            ws4.cell(row=i+1, column=2, value=sug)
        ws4.column_dimensions['A'].width = 8
        ws4.column_dimensions['B'].width = 80

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer


export_service = ExportService()
